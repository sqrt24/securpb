from datetime import date
from io import BytesIO

import openpyxl
from django.http import HttpResponse
from django.template.response import TemplateResponse

from blog.models import Article, ArticleCategory
from services.models import PriceListCategory, PriceListItem, Service, ServiceCategory


def robots_view(request):
    return HttpResponse(
        request.project_settings.robots_txt
        + f"\n\nSitemap: https://{request.get_host()}/sitemap.xml"
        + f"\nHost: https://{request.get_host()}",
        content_type="text/plain",
        charset="utf-8",
        headers={"Content-Encoding": "utf-8"},
    )


def sitemap_view(request):
    return TemplateResponse(
        request,
        "seo/sitemap.xml",
        {
            "lastmod": str(date.today()),
            "service_categories": ServiceCategory.objects.all(),
            "services": Service.objects.filter(is_active=True),
            "articles": Article.objects.filter(is_active=True),
            "article_categories": ArticleCategory.objects.all(),
        },
        content_type="application/xml",
    )


def feed_view(request):
    return TemplateResponse(
        request,
        "seo/feed.xml",
        {
            "categories": PriceListCategory.objects.all(),
            "items": PriceListItem.objects.all().select_related("category"),
        },
        content_type="application/xml",
    )


def service_feed_view(request):
    return TemplateResponse(
        request,
        "seo/service-feed.xml",
        {
            "categories": ServiceCategory.objects.all(),
            "items": Service.objects.all().select_related("category"),
        },
        content_type="application/xml",
    )


def xlsx_2gis_view(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(1, 1, "Наименование товара")
    worksheet.cell(1, 2, "Цена")
    worksheet.cell(1, 3, "Категория")
    worksheet.cell(1, 4, "Ссылка на товар на сайте магазина")
    worksheet.cell(1, 5, "Ссылка на картинку")
    worksheet.cell(1, 6, "Описание")

    for row_idx, item in enumerate(PriceListItem.objects.all().select_related("category"), 2):
        worksheet.cell(row_idx, 1, f"{item.category.name}: {item.name}")
        worksheet.cell(row_idx, 2, item.integer_price)
        worksheet.cell(row_idx, 3, item.category.name)
        worksheet.cell(row_idx, 4, "https://securpb.ru/price-list/")
        worksheet.cell(row_idx, 5, "")
        worksheet.cell(row_idx, 6, "")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.ms-excel",
        headers={"Content-Disposition": f"attachment; filename={date.today()}.xlsx"},
    )
